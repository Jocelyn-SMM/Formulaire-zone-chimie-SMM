from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def envoyer_notification(donnees):
    # --- CONFIGURATION DU SERVEUR MAIL ---
    smtp_server = "smtp.office365.com"
    smtp_port = 587
    expediteur = "jcoisine@smm-composites.com"
    mot_de_passe = "*jojo83*"
    
    # Liste des personnes qui vont recevoir le mail
    destinataires = ["jcoisine@smm-composites.com"]

    # --- CRÉATION DU MESSAGE ---
    msg = MIMEMultipart()
    msg['From'] = expediteur
    msg['To'] = ", ".join(destinataires)
    msg['Subject'] = f"🔔 Nouvel Audit Chimie réalisé par {donnees['Auditeur']} 🔔"

    # Corps du mail
    corps_mail = f"""
    Bonjour,

    Un nouvel audit de la zone chimie vient d'être validé.
    
    Détails principaux :
    - Date : {donnees['Date']}
    - Auditeur : {donnees['Auditeur']}
    - Remarques : {donnees['Remarques']}

    Veuillez consulter le fichier Excel audits sur le réseau (N:\05-HSE\9 - Audits de zone\Application audit de zone\Programmation) pour voir l'audit complet avec les photos et la signature.

    Ceci est un message automatique.
    """
    msg.attach(MIMEText(corps_mail, 'plain', 'utf-8'))

    # --- ENVOI ---
    try:
        serveur = smtplib.SMTP(smtp_server, smtp_port)
        serveur.starttls() # Sécurise la connexion
        serveur.login(expediteur, mot_de_passe)
        serveur.sendmail(expediteur, destinataires, msg.as_string())
        serveur.quit()
        print("✉️ E-mail de notification envoyé avec succès !")
    except Exception as e:
        print(f"⚠️ Erreur lors de l'envoi de l'e-mail : {e}")

app = Flask(__name__)
CORS(app)

EXCEL_FILE = "audits.xlsx"
UPLOAD_FOLDER = "images_audits"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/api/audit', methods=['POST'])
def recevoir_audit():
    try:
        # 1. Extraction des textes et choix Oui/Non
        donnees = {
            "Date": request.form.get("date_audit"),
            "Auditeur": request.form.get("nom_auditeur"),
            "Zone balayée": request.form.get("zone_balayee"),
            "Déchets": request.form.get("dechets"),
            "Produit au sol": request.form.get("produit_sol"),
            "Carton Epoxy": request.form.get("epoxy_change"),
            "Pourquoi pas changé": request.form.get("pourquoi_epoxy", ""),
            "DLU": request.form.get("dlu_conforme"),
            "Compatibilité fûts": request.form.get("compatibilite_futs"),
            "Action fûts": request.form.get("action_futs", ""),
            "Entretien Balances": request.form.get("entretien_balances"),
            "Remarques": request.form.get("remarques_semaine", "")
        }

        # Liste exacte de tous les fichiers attendus
        liste_fichiers = ['photo_résultat', 'photo_epoxy', 'photo_balances', 'signature']

        # 2. Gestion et sauvegarde des fichiers sur le PC
        chemins_fichiers = {}
        for clé in liste_fichiers:
            fichier = request.files.get(clé)
            if fichier:
                nom_fichier = f"{donnees['Auditeur']}_{donnees['Date']}_{clé}_{fichier.filename}".replace("/", "-").replace(" ", "_")
                chemin_complet = os.path.join(UPLOAD_FOLDER, nom_fichier)
                fichier.save(chemin_complet)
                chemins_fichiers[clé] = chemin_complet
            else:
                chemins_fichiers[clé] = None

        # --- 3. Gestion intelligente du fichier Excel ---
        if os.path.exists(EXCEL_FILE):
            # Le fichier existe déjà : on tente de le charger
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
                ws = wb.active
                print("Fichier Excel existant chargé avec succès.")
            except PermissionError:
                # Sécurité si le fichier est resté ouvert dans Excel sur ton PC
                return jsonify({"status": "error", "message": "⚠️ Ferme le fichier audits.xlsx sur ton ordinateur avant de valider !"}), 500
        else:
            # Le fichier n'existe pas du tout : on le crée à neuf avec ses entêtes
            wb = Workbook()
            ws = wb.active
            ws.title = "Audits"
            entêtes = list(donnees.keys()) + ["Photo Résultat", "Photo Carton Époxy", "Photo Balances", "Signature"]
            ws.append(entêtes)
            wb.save(EXCEL_FILE)
            print("Nouveau fichier Excel créé avec les entêtes.")

        # --- 4. Ajout des données textuelles sur la ligne suivante ---
        prochaine_ligne = ws.max_row + 1
        valeurs_textes = list(donnees.values())
        
        for col_idx, valeur in enumerate(valeurs_textes, start=1):
            ws.cell(row=prochaine_ligne, column=col_idx, value=valeur)

        # On agrandit la hauteur de la ligne pour que les images ne bavent pas
        ws.row_dimensions[prochaine_ligne].height = 80

        # --- 5. Insertion visuelle des 4 images avec traitement de transparence ---
        # On cale les colonnes pour correspondre exactement à l'ordre des en-têtes :
        # K = Photo Principale (13), L = Photo Carton Époxy (7), M = Photo Balances (11), N = Signature (14)
        colonnes_images = {
            'photo_résultat': 13,
            'photo_epoxy': 14,
            'photo_balances': 15,
            'signature': 16
        }

        for clé, col_idx in colonnes_images.items():
            chemin = chemins_fichiers.get(clé)
            lettre_colonne = openpyxl.utils.get_column_letter(col_idx)
            
            if chemin is not None and os.path.exists(chemin):
                try:
                    # Traitement de l'image avec Pillow pour supprimer le fond transparent noir
                    with PILImage.open(chemin) as pil_img:
                        # Si l'image a une transparence (comme le dessin de la signature), on lui met un fond blanc
                        if pil_img.mode in ('RGBA', 'LA') or (pil_img.mode == 'P' and 'transparency' in pil_img.info):
                            fond_blanc = PILImage.new('RGBA', pil_img.size, (255, 255, 255))
                            fond_blanc.paste(pil_img, mask=pil_img.split()[3] if pil_img.mode == 'RGBA' else None)
                            rgb_img = fond_blanc.convert('RGB')
                        else:
                            rgb_img = pil_img.convert('RGB')
                        
                        rgb_img.save(chemin) # On réenregistre l'image propre

                    # Ajout dans la cellule Excel
                    img = OpenpyxlImage(chemin)
                    img.width = 100
                    img.height = 75
                    
                    cellule_cible = f"{lettre_colonne}{prochaine_ligne}"
                    ws.add_image(img, cellule_cible)
                    ws.column_dimensions[lettre_colonne].width = 18
                except Exception as img_err:
                    print(f"Erreur traitement {clé}: {img_err}")
                    ws.cell(row=prochaine_ligne, column=col_idx, value="Format non supporté")
            else:
                # Si aucune image n'est fournie (ex: Balances "Aucune"), on écrit proprement le texte
                ws.cell(row=prochaine_ligne, column=col_idx, value="Aucune")

        wb.save(EXCEL_FILE)
        try:
            envoyer_notification(donnees)
        except Exception as mail_err:
            print(f"L'envoi du mail a échoué, mais l'audit est sauvé : {mail_err}")
        
        return jsonify({"status": "success", "message": "Audit validé avec succès !"}), 200

    except Exception as e:
        # Ce bloc-ci ne s'exécute QUE si le code plante gravement
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == '__main__':
    app.run(port=5000, debug=True)
